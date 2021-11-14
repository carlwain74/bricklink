"""
This is neat little utility that gets pricing for Lego sets.
"""
import json
import argparse
import sys
import logging
import os
from os import stat
from os.path import exists
from bricklink_api.auth import oauth
from bricklink_api.catalog_item import get_price_guide, get_item, get_item_image, Type, NewOrUsed
from bricklink_api.category import get_category
import html
from html.parser import HTMLParser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment,Font,PatternFill
import configparser
from datetime import datetime

logging.basicConfig(
    format='%(asctime)s %(levelname)-8s %(message)s',
    level=logging.INFO,
    datefmt='%Y-%m-%d %H:%M:%S')

"""
This calls the API functions to get the data.
"""
def getDetails(set_number, auth_params):
    logging.debug("Getting details for " + str(set_number))
    h_parse = html.parser

    if set_number == "40158":
        item_type = Type.GEAR
    else:
        item_type = Type.SET

    json_obj = get_price_guide(item_type, set_number, new_or_used=NewOrUsed.NEW, \
                               country_code="US", region="north_america", auth=auth_params)

    logging.debug(json.dumps(json_obj, indent=4, sort_keys=True))
    meta = json_obj['meta']

    if meta['code'] == 200:
        data = json_obj['data']

        type_data = get_item(item_type, set_number, auth=auth_params)
        logging.debug(json.dumps(type_data, indent=4, sort_keys=True))

        category_data = get_category(type_data['data']['category_id'], auth=auth_params)
        logging.debug(json.dumps(category_data, indent=4, sort_keys=True))

        elem_data = {}
        elem_data[set_number] = {}
        elem_data[set_number]['name'] = h_parse.unescape(type_data['data']['name'])
        elem_data[set_number]['category'] = h_parse.unescape(category_data['data']['category_name'])
        elem_data[set_number]['avg'] = round(int(float(data['avg_price'])))
        elem_data[set_number]['max'] = round(int(float(data['max_price'])))
        elem_data[set_number]['min'] = round(int(float(data['min_price'])))
        elem_data[set_number]['quantity'] = data['unit_quantity']
        elem_data[set_number]['currency'] = data['currency_code']

        return elem_data
    else:
        logging.warning("API Error!! " + str(meta['code']))
        logging.warning("API Message!! " + str(meta['message']))
        return 0

"""
This prints stuff to the screen.
"""
def print_details(element_data, number):
    logging.info("Item: " + number)
    logging.debug("  Name: " + element_data['name'])
    logging.debug("  Category: " + element_data['category'])
    logging.debug("  Avg Price: " + str(element_data['avg']) + " " + element_data['currency'])
    logging.debug("  Max Price: " + str(element_data['max']) + " " + element_data['currency'])
    logging.debug("  Min Price: " + str(element_data['min']) + " " + element_data['currency'])
    logging.debug("  Quantity avail: " + str(element_data['quantity']))

"""
Create workbook
"""
def setup_xls_writer(xls_filename):
    try:
        if os.path.isfile(xls_filename) and os.access(xls_filename, os.R_OK):
            logging.info('Load excel file')
            workbook = load_workbook(filename=xls_filename)
        else:
            workbook = Workbook()
            logging.info(workbook.sheetnames)
            workbook.remove(workbook['Sheet'])
    except Exception as exception:
        logging.error('Could not load excel file!' + str(exception))
        sys.exit(1)

    now = datetime.now() # current date and time
    date_stamp = now.strftime("%m_%d_%Y__%H%M")
    return workbook

"""
Add workbook unless it already exists
"""
def add_worksheet(workbook, item_name):
    # See if the worksheet already exists
    if item_name in workbook.sheetnames:
        worksheet = workbook[item_name]
    else:
        worksheet = workbook.create_sheet(item_name, 0)

        # Start from the first cell. Rows and columns are zero indexed.
        _row = 2
        _col = 2

        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20

        header_color = "00C0C0C0"
        d = worksheet.cell(row=2, column=2, value="Name")
        d.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        d.alignment = Alignment(horizontal="center", vertical="center")

        d = worksheet.cell(row=3, column=2, value="Category")
        d.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        d.alignment = Alignment(horizontal="center", vertical="center")
        
        xls_headers = ['Date', 'Avg Price', 'Min Price', 'Max Price', 'Quantity']

        _row = 5
        col_adjust = 0
        for headers in xls_headers:
            #worksheet.write(row, col+col_adjust, headers, header_format)
            d = worksheet.cell(row=_row, column=_col+col_adjust, value=headers)
            d.alignment = Alignment(horizontal="center", vertical="center")
            d.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            col_adjust += 1

    return worksheet

"""
The main routine.
"""
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--set', type=str)
    parser.add_argument('-f', '--file', type=str)
    parser.add_argument('-v', '--verbose', action="store_true")
    args = parser.parse_args()

    set_num = args.set
    filename = args.file

    logging.info('Read configuration')
    config = configparser.ConfigParser()
    config.read('config.ini')

    # fill in with your data from https://www.bricklink.com/v2/api/register_consumer.page
    consumer_key = config['secrets']['consumer_key']
    consumer_secret = config['secrets']['consumer_secret']
    token_value = config['secrets']['token_value']
    token_secret = config['secrets']['token_secret']
    try:
        auth = oauth(consumer_key, consumer_secret, token_value, token_secret)
    except Exception as e:
        logging.error('Could not get auth token')
        sys.exit(1)

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if set_num:
        logging.info('Processing single set')
        res = getDetails(set_num, auth)
        if not res:
            sys.exit(1)
        logging.debug(json.dumps(res, indent=4, sort_keys=True))
        for key in res:
            print_details(res[key], key)
    elif filename:
        xls_filename = 'Items.xlsx'
        workbook = setup_xls_writer(xls_filename)

        logging.info('Processing multiple sets')
        if exists(filename):
            logging.info("Processing sets in " + filename)

            if stat(filename).st_size == 0:
                logging.error("File is empty!!")
                sys.exit()
            else:
                file_handler = open(filename, "r")
                total = 0
                _row = 6
                _col = 2
                now = datetime.now()
                date_stamp = now.strftime("%m-%d-%Y")
                while True:
                    line = file_handler.readline()
                    if not line:
                        break
                    #print(line.strip())
                    number = line.strip()
                    res = getDetails(number, auth)
                    if not res:
                        sys.exit(1)
                    for key in res:
                        worksheet = add_worksheet(workbook, key)

                        # Find next available row on column B
                        for index in range(6, 1000):
                            if worksheet.cell(row=index, column=2).value is None:
                                _row = index
                                logging.debug('Inserting at ros ' + str(_row))
                                break
                            else:
                                logging.debug('Row contents: '+
                                    worksheet.cell(row=index, column=2).value)

                        print_details(res[key], key)
                        logging.debug(json.dumps(res, indent=4, sort_keys=True))
                        total += res[key]['avg']

                        d = worksheet.cell(row=2, column=3, value=res[key]['name'])
                        d.alignment = Alignment(horizontal="center", vertical="center")
                        d = worksheet.cell(row=3, column=3, value=res[key]['category'])
                        d.alignment = Alignment(horizontal="center", vertical="center")
                        d = worksheet.cell(row=_row, column=_col, value=date_stamp)
                        d = worksheet.cell(row=_row, column=_col+1, value=res[key]['avg'])
                        d.alignment = Alignment(horizontal="center", vertical="center")
                        d = worksheet.cell(row=_row, column=_col+2, value=res[key]['min'])
                        d.alignment = Alignment(horizontal="center", vertical="center")
                        d = worksheet.cell(row=_row, column=_col+3, value=res[key]['max'])
                        d.alignment = Alignment(horizontal="center", vertical="center")
                        d = worksheet.cell(row=_row, column=_col+4, value=res[key]['quantity'])
                        d.alignment = Alignment(horizontal="center", vertical="center")

                logging.info("Total: " + str(total) + "USD")

                if 'Summary' in workbook.sheetnames:
                    summary = workbook['Summary']
                else:
                    summary = workbook.create_sheet("Summary", 0)

                    summary.column_dimensions['B'].width = 10
                    summary.column_dimensions['C'].width = 20

                    header_color = "00C0C0C0"
                    d = summary.cell(row=2, column=2, value="Date")
                    d.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                    d.alignment = Alignment(horizontal="center", vertical="center")

                    d = summary.cell(row=2, column=3, value="Total")
                    d.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                    d.alignment = Alignment(horizontal="center", vertical="center")

                for index in range(3, 1000):
                    if summary.cell(row=index, column=2).value is None:
                        _srow = index
                        logging.debug('Inserting at ros ' + str(_srow))
                        break
                    else:
                        logging.debug('Row contents: '+summary.cell(row=index, column=2).value)

                d = summary.cell(row=_srow, column=2, value=date_stamp)
                d.alignment = Alignment(horizontal="center", vertical="center")
                d = summary.cell(row=_srow, column=3, value=total)
                d.alignment = Alignment(horizontal="center", vertical="center")
        workbook.save(filename=xls_filename)

if __name__ == '__main__':
    main()
