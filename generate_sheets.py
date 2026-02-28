"""
This is neat little utility that gets pricing for Lego sets.
"""
import json
import argparse
import sys
import logging
import os
from os import stat
from os.path import exists
from bricklink_py import Bricklink
import html
from html.parser import HTMLParser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment,Font,PatternFill
import configparser
from datetime import datetime

logging.basicConfig(
format='%(asctime)s %(levelname)-8s %(message)s',
level=logging.DEBUG,
datefmt='%Y-%m-%d %H:%M:%S')


def get_last_sale_date(sales: dict) -> str | None:
    """
    Given an unordered dictionary of past sales (keyed by any value),
    return the ISO 8601 date string of the most recent sale.

    Each sale entry is expected to contain a 'date_ordered' field in
    ISO 8601 format, e.g. '2023-05-27T01:09:39.493Z'.

    Returns the raw ISO string of the most recent sale, or None if the
    dictionary is empty or no valid dates are found.

    Example
    -------
    sales = {
        1: {"date_ordered": "2023-05-27T01:09:39.493Z", "unit_price": "197.42"},
        2: {"date_ordered": "2023-12-11T18:44:02.100Z", "unit_price": "210.00"},
        3: {"date_ordered": "2022-08-03T09:15:55.000Z", "unit_price": "185.00"},
    }
    get_last_sale_date(sales)
    # → '2023-12-11T18:44:02.100Z'
    """
    latest_dt  = None
    latest_raw = None

    items = sales.values() if isinstance(sales, dict) else sales
    for sale in items:
        raw = sale.get('date_ordered', '')
        if not raw:
            continue
        try:
            # Replace trailing Z with +00:00 for fromisoformat compatibility
            dt = datetime.fromisoformat(raw.replace('Z', '+00:00'))
        except ValueError:
            continue

        if latest_dt is None or dt > latest_dt:
            latest_dt  = dt
            latest_raw = raw

    return latest_raw

"""
This calls the API functions to get the data.
"""
def getDetails(session, set_number):
    logging.debug("Getting details for " + str(set_number))
    h_parse = html.parser

    if set_number == "40158":
        item_type = "GEAR"
    else:
        item_type = "SET"

    try:
        current_items = session.catalog_item.get_price_guide(item_type, set_number, new_or_used="N",
                                                             country_code="US", region="north_america")

        past_sales = session.catalog_item.get_price_guide(item_type, set_number, new_or_used="N", \
                                                          guide_type="sold", country_code="US", region="north_america")
    except Exception as e:
        logging.exception("Failed to get price guide for item" + str(e))
        return {}

    logging.debug(json.dumps(current_items, indent=4, sort_keys=True))
    logging.debug(json.dumps(past_sales, indent=4, sort_keys=True))

    type_data = session.catalog_item.get_item(item_type, set_number)

    logging.debug(json.dumps(type_data, indent=4, sort_keys=True))

    category_data = session.category.get_category(type_data['category_id'])
    logging.debug(json.dumps(category_data, indent=4, sort_keys=True))

    elem_data = {}
    elem_data[set_number] = {}
    elem_data[set_number]['name'] = h_parse.unescape(type_data['name'])
    elem_data[set_number]['category'] = h_parse.unescape(category_data['category_name'])
    elem_data[set_number]['current'] = {}
    elem_data[set_number]['current']['avg'] = round(int(float(current_items['avg_price'])))
    elem_data[set_number]['current']['max'] = round(int(float(current_items['max_price'])))
    elem_data[set_number]['current']['min'] = round(int(float(current_items['min_price'])))
    elem_data[set_number]['current']['quantity'] = current_items['unit_quantity']
    elem_data[set_number]['current']['currency'] = current_items['currency_code']
    elem_data[set_number]['past'] = {}
    elem_data[set_number]['past']['avg'] = round(int(float(past_sales['avg_price'])))
    elem_data[set_number]['past']['max'] = round(int(float(past_sales['max_price'])))
    elem_data[set_number]['past']['min'] = round(int(float(past_sales['min_price'])))
    elem_data[set_number]['past']['quantity'] = past_sales['unit_quantity']
    elem_data[set_number]['past']['currency'] = past_sales['currency_code']
    elem_data[set_number]['past']['last_sale_date'] = get_last_sale_date(past_sales['price_detail'])
    elem_data[set_number]['year'] = type_data['year_released']
    elem_data[set_number]['image'] = type_data['image_url']
    elem_data[set_number]['thumbnail'] = type_data['thumbnail_url']


    return elem_data

"""
This prints stuff to the screen.
"""
def print_details(element_data, number):
    logging.info("Item: " + number)
    logging.info("  Name: " + element_data['name'])
    logging.info("  Category: " + element_data['category'])
    logging.info("  Current Sales: ")
    logging.info("     Average: " + str(element_data['current']['avg']) + " " + element_data['current']['currency'])
    logging.info("     Max: " + str(element_data['current']['max']) + " " + element_data['current']['currency'])
    logging.info("     Min: " + str(element_data['current']['min']) + " " + element_data['current']['currency'])
    logging.info("     Quantity avail: " + str(element_data['current']['quantity']))
    logging.info("  Previous Sales: ")
    logging.info("     Average: " + str(element_data['past']['avg']) + " " + element_data['past']['currency'])
    logging.info("     Max: " + str(element_data['past']['max']) + " " + element_data['past']['currency'])
    logging.info("     Min: " + str(element_data['past']['min']) + " " + element_data['past']['currency'])
    logging.info("     Quantity avail: " + str(element_data['past']['quantity']))
    logging.info("     Last Sale Date: " + str(element_data['past']['last_sale_date']))
    logging.info("  Year Released: " + str(element_data['year']))
    logging.info("  Image: " + str(element_data['image']))
    logging.info("  Thumbnail: " + str(element_data['thumbnail']))

"""
Create workbook
"""
def create_wookbook(xls_filename):
    try:
        if os.path.isfile(xls_filename) and os.access(xls_filename, os.R_OK):
            logging.info('Load excel file')
            workbook = load_workbook(filename=xls_filename)
        else:
            workbook = Workbook()
            logging.info(workbook.sheetnames)
            workbook.remove(workbook['Sheet'])
    except Exception as exception:
        logging.error('Could not load excel file!' + str(exception))
        sys.exit(1)

    return workbook

"""
Add workbook unless it already exists
"""
def add_worksheet(workbook, item_name):
    # See if the worksheet already exists
    if item_name in workbook.sheetnames:
        worksheet = workbook[item_name]
    else:
        worksheet = workbook.create_sheet(item_name, 0)

        # Start from the first cell. Rows and columns are zero indexed.
        _row = 2
        _col = 2

        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20

        header_color = "00C0C0C0"
        data = worksheet.cell(row=2, column=2, value="Name")
        data.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

        data = worksheet.cell(row=3, column=2, value="Category")
        data.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

        xls_headers = ['Date', 'Avg Price', 'Min Price', 'Max Price', 'Quantity']

        _row = 5
        col_adjust = 0
        for headers in xls_headers:
            #worksheet.write(row, col+col_adjust, headers, header_format)
            data = worksheet.cell(row=_row, column=_col+col_adjust, value=headers)
            data.alignment = Alignment(horizontal="center", vertical="center")
            data.fill = PatternFill(start_color=header_color,
                                    end_color=header_color, fill_type="solid")
            col_adjust += 1

    return worksheet

def create_wookbook_and_sheet(xls_filename):
    workbook = create_wookbook(xls_filename)

    now = datetime.now() # current date and time
    date_stamp = now.strftime("%m_%d_%Y")
    worksheet = workbook.create_sheet('Items_'+date_stamp)

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 1

    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['F'].width = 20

    header_color = "00C0C0C0"

    xls_headers = ['Item', 'Name', 'Category', 'Avg Price', 'Min Price', 'Max Price', 'Quantity', 'Year']

    _row = 5
    col_adjust = 0
    for headers in xls_headers:
        #worksheet.write(row, col+col_adjust, headers, header_format)
        data = worksheet.cell(row=row, column=col+col_adjust, value=headers)
        data.alignment = Alignment(horizontal="center", vertical="center")
        data.fill = PatternFill(start_color=header_color,
                                end_color=header_color, fill_type="solid")
        col_adjust += 1

    return workbook, worksheet


def create_api_session(config_file):

    config = configparser.ConfigParser()
    config.read(config_file)

    # fill in with your data from https://www.bricklink.com/v2/api/register_consumer.page
    consumer_key = config['secrets']['consumer_key']
    consumer_secret = config['secrets']['consumer_secret']
    token_value = config['secrets']['token_value']
    token_secret = config['secrets']['token_secret']

    try:
        session = Bricklink(
            consumer_key=consumer_key,
            consumer_secret=consumer_secret,
            token=token_value,
            token_secret=token_secret
        )
    except Exception as e:
        logging.error('Could not get auth token - ', str(e))
        return null

    return session


def generate_single_sheet(session, file_handler, workbook, worksheet):
    logging.info('Writing all sets to the same file')
    total = 0
    _row = 1
    _col = 1
    while True:
        line = file_handler.readline()
        if not line:
            break
        #print(line.strip())
        number = line.strip()
        res = getDetails(session, number)
        if not res:
            sys.exit(1)
        for key in res:
            print_details(res[key], key)
            logging.debug(json.dumps(res, indent=4, sort_keys=True))
            total += res[key]['current']['avg']

            _row += 1

            data = worksheet.cell(row=_row, column=_col, value=key)
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+1, value=res[key]['name'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+2, value=res[key]['category'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+3, value=res[key]['current']['avg'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+4, value=res[key]['current']['min'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+5, value=res[key]['current']['max'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+6, value=res[key]['current']['quantity'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+7, value=res[key]['year'])
            data.alignment = Alignment(horizontal="center", vertical="center")

    logging.info("Total: " + str(total) + "USD")

def generate_multi_sheet(session, file_handler, workbook):

    logging.info("Writing sets per sheet`")

    total = 0
    _row = 6
    _col = 2

    now = datetime.now()
    date_stamp = now.strftime("%m-%d-%Y")

    while True:
        line = file_handler.readline()
        if not line:
            break
        number = line.strip()
        res = getDetails(session, number)
        if not res:
            logging.error('Could not get details for set:' +number)
        for key in res:
            worksheet = add_worksheet(workbook, key)
            # Find next available row on column B
            for index in range(6, 1000):
                if worksheet.cell(row=index, column=2).value is None:
                    _row = index
                    logging.debug('Inserting at ros ' + str(_row))
                    break
                else:
                    logging.debug('Row contents: '+
                                  worksheet.cell(row=index, column=2).value)

            print_details(res[key], key)
            logging.debug(json.dumps(res, indent=4, sort_keys=True))
            total += res[key]['avg']

            data = worksheet.cell(row=2, column=3, value=res[key]['name'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=3, column=3, value=res[key]['category'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col, value=date_stamp)
            data = worksheet.cell(row=_row, column=_col+1, value=res[key]['current']['avg'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+2, value=res[key]['current']['min'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+3, value=res[key]['current']['max'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+4, value=res[key]['current']['quantity'])
            data.alignment = Alignment(horizontal="center", vertical="center")

    logging.info("Total: " + str(total) + "USD")

    if 'Summary' in workbook.sheetnames:
        summary = workbook['Summary']
    else:
        summary = workbook.create_sheet("Summary", 0)

        summary.column_dimensions['B'].width = 10
        summary.column_dimensions['C'].width = 20

        header_color = "00C0C0C0"
        data = summary.cell(row=2, column=2, value="Date")
        data.fill = PatternFill(start_color=header_color,
                                end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

        data = summary.cell(row=2, column=3, value="Total")
        data.fill = PatternFill(start_color=header_color,
                                end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

    for index in range(3, 1000):
        if summary.cell(row=index, column=2).value is None:
            _srow = index
            logging.debug('Inserting at ros ' + str(_srow))
            break
        else:
            logging.debug('Row contents: '+summary.cell(row=index, column=2).value)

    data = summary.cell(row=_srow, column=2, value=date_stamp)
    data.alignment = Alignment(horizontal="center", vertical="center")
    data = summary.cell(row=_srow, column=3, value=total)
    data.alignment = Alignment(horizontal="center", vertical="center")

def test_config(config_file = 'config.ini'):
    session = create_api_session(config_file)
    res = getDetails(session, "75105-1")

    if res:
        return True
    else:
        return False

"""
The main handler routine.
"""
def sheet_handler(set_num, set_list, multi_sheet, output_file = 'Sets.xlsx', config_file = 'config.ini'):
    
    logging.info('Setup API session')
    session = create_api_session(config_file)

    if not session:
        logging.error('Could not create an API session')
        sys.exit(1)

    if set_num:
        logging.info('Processing single set')
        try:
            res = getDetails(session, set_num)
        except Exception as e:
            logging.exception("Could not get set details" + str(e))
            return None

        logging.debug(json.dumps(res, indent=4, sort_keys=True))
        for key in res:
            print_details(res[key], key)
    elif set_list:
        xls_filename = output_file

        if multi_sheet:
            workbook = create_wookbook(xls_filename)
        else:
            (workbook, worksheet) = create_wookbook_and_sheet(xls_filename)

        logging.info('Processing multiple sets')
        if exists(set_list):
            logging.info("Processing sets in " + set_list)

            if stat(set_list).st_size == 0:
                logging.error("File is empty!!")
                sys.exit()
            else:
                file_handler = open(set_list, "r")
                now = datetime.now()
                date_stamp = now.strftime("%m-%d-%Y")

                # Sheet per item and Summary
                if multi_sheet:
                    generate_multi_sheet(session, file_handler, workbook)
                else:
                    generate_single_sheet(session, file_handler, workbook, worksheet)

            workbook.save(filename=xls_filename)

if __name__ == '__main__':
    sheet_handler("71016-1", "", False, False)
