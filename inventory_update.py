"""
This is neat little utility that updates inventory on Bricklink.
"""
import json
import argparse
import sys
import logging
import os
from os import stat
from os.path import exists
from bricklink_api.auth import oauth
from bricklink_api.catalog_item import get_price_guide, get_item, get_item_image, Type, NewOrUsed
from bricklink_api.category import get_category
from bricklink_api.user_inventory import get_inventory, create_inventory, update_inventory
from bricklink_api.color import get_color_list
import html
from html.parser import HTMLParser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment,Font,PatternFill
import configparser
from datetime import datetime


logging.basicConfig(
    format='%(asctime)s %(levelname)-8s %(message)s',
    level=logging.INFO,
    datefmt='%Y-%m-%d %H:%M:%S')

"""
Grab Color from loopup table
"""
def getColorName(colorId):
    pass

"""
This calls the API functions to get the data.
"""
def getPartDetails(number, auth_params):
    logging.debug("Getting details for " + str(number))
    h_parse = html.parser

    json_obj = get_price_guide(Type.PART, number, new_or_used=NewOrUsed.USED, \
                               country_code="US", region="north_america", auth=auth_params)

    logging.debug(json.dumps(json_obj, indent=4, sort_keys=True))
    meta = json_obj['meta']

    if meta['code'] == 200:
        data = json_obj['data']

        type_data = get_item(Type.PART, number, auth=auth_params)
        logging.debug(json.dumps(type_data, indent=4, sort_keys=True))

        category_data = get_category(type_data['data']['category_id'], auth=auth_params)
        logging.debug(json.dumps(category_data, indent=4, sort_keys=True))

        elem_data = {}
        elem_data[number] = {}
        elem_data[number]['name'] = h_parse.unescape(type_data['data']['name'])
        elem_data[number]['category'] = h_parse.unescape(category_data['data']['category_name'])
        elem_data[number]['avg'] = float(data['avg_price'])
        elem_data[number]['max'] = float(data['max_price'])
        elem_data[number]['min'] = float(data['min_price'])
        elem_data[number]['quantity'] = data['unit_quantity']
        elem_data[number]['currency'] = data['currency_code']

        return elem_data
    else:
        logging.warning("API Error!! " + str(meta['code']))
        logging.warning("API Message!! " + str(meta['message']))
        return 0

"""
Open inventory workbook
"""
def setup_xls_writer(xls_filename):
    try:
        if os.path.isfile(xls_filename) and os.access(xls_filename, os.R_OK):
            logging.info('Load excel file')
            workbook = load_workbook(filename=xls_filename)
    except Exception as exception:
        logging.error('Could not load excel file!' + str(exception))
        sys.exit(1)

    return workbook

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('-v', '--verbose', action="store_true")
    parser.add_argument('-s', '--skip', action="store_true")
    parser.add_argument('-d', '--dryrun', action="store_true")
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    logging.info('Read configuration')
    config = configparser.ConfigParser()
    config.read('config.ini')

    # Read color conversion data
    try:
        with open('colors.json') as configData:
            configData = json.load(configData)
            # Process dict
            logging.debug(json.dumps(configData, indent=4, sort_keys=True))
    except IOError as e:
        logging.critical("Could not open: " + args.config)
        sys.exit(2)
    
    # fill in with your data from https://www.bricklink.com/v2/api/register_consumer.page
    consumer_key = config['secrets']['consumer_key']
    consumer_secret = config['secrets']['consumer_secret']
    token_value = config['secrets']['token_value']
    token_secret = config['secrets']['token_secret']

    try:
        auth = oauth(consumer_key, consumer_secret, token_value, token_secret)
    except Exception as error:
        logging.error('Could not get auth token' + str(error))
        sys.exit(1)

    workbook = setup_xls_writer('LegoParts.xlsx')

    worksheet = workbook['Inventory']
    index = 4
    while worksheet.cell(row=index, column=3).value is not None:
        logging.debug('Column Index: ' + str(index))
        
        # Check if there is an inventory id
        if worksheet.cell(row=index, column=2).value is not None:
            inventory_id = worksheet.cell(row=index, column=2).value
            logging.debug("Entry has Inventory Id: " + str(inventory_id))
            
            if args.skip:
                index += 1
                continue
        else:
            inventory_id = 0

        item_type = worksheet.cell(row=index, column=3).value
        logging.info('')
        logging.info('Processing: '+ str(item_type))
        if inventory_id:
            logging.info("  Inventory Id: " + str(inventory_id))
        item_num = worksheet.cell(row=index, column=4).value
        logging.info("  Item Num: " + str(item_num))
        color = worksheet.cell(row=index, column=6).value
        logging.info("  Color: " + str(color))
        price = worksheet.cell(row=index, column=7).value
        logging.info("  Price: " + str(price))
        quantity = worksheet.cell(row=index, column=8).value
        logging.info("  Quantity: " + str(quantity))
        condition = worksheet.cell(row=index, column=9).value
        logging.debug("  Condition: " + str(condition))
        description = worksheet.cell(row=index, column=12).value
        logging.debug("  Description: " + str(description))
        stockroom = worksheet.cell(row=index, column=15).value
        logging.debug("  Stockroom: " + str(stockroom))
        remark = worksheet.cell(row=index, column=14).value
        logging.debug("  Remark: " + str(remark))
        stockroom_id = worksheet.cell(row=index, column=16).value
        logging.debug("  Stockroom Id: " + str(stockroom_id))
        retain = worksheet.cell(row=index, column=17).value
        logging.debug("  Stockroom Id: " + str(retain))

        if item_num is None:
            index += 1
            logging.info('Empty row!!')
            continue

        if quantity == 0 or quantity is None:
            logging.warning('No quantity provided or it\'s zero')
            index += 1
            continue

        if color == 0 or color is None:
            logging.warning('No color provided or it\'s zero')
            index += 1
            continue
            
        # Call Bricklink API
        inventory_item = {}
        inventory_item['item'] = {}
        inventory_item['item']['type'] = item_type
        inventory_item['item']['no'] = item_num
        inventory_item['color_id'] = color
        inventory_item['unit_price'] = price
        
        inventory_item['new_or_used'] = condition
        inventory_item['description'] = description
        inventory_item['is_stock_room'] = stockroom
        inventory_item['stock_room_id'] = stockroom_id
        inventory_item['is_retain'] = retain
        inventory_item['remarks'] = configData[str(color)]['Name']
        logging.debug(json.dumps(inventory_item, indent=4, sort_keys=True))

        if inventory_id == 0:
            logging.info('Creating Inventory Item')
            # Get price details
            try:
                details = getPartDetails(item_num, auth)
                if price is None:
                    logging.debug(details)
                    inventory_item['unit_price'] = details[item_num]['avg']
                    if not args.dryrun:
                        worksheet.cell(row=index, column=7).value = details[item_num]['avg']
                    else:
                        logging.info('  Avg Unit Price: ' + str(inventory_item['unit_price']))
                if not args.dryrun:
                    worksheet.cell(row=index, column=24).value = details[item_num]['name']
                inventory_item['quantity'] = quantity
            except Exception as e:
                logging.warning('Could not get pricing details for ' + str(item_num))
                logging.warning(details)
                index += 1
                continue
            if not args.dryrun:
                try:
                    response = create_inventory(inventory_item, auth=auth)
                    logging.debug(response)
                    inventory_id = response['data']['inventory_id']
                    unit_price = response['data']['unit_price']
                    logging.info('  Inventory Id: ' + str(inventory_id))
                    logging.info('  Avg Unit Price: ' + str(unit_price))
                    worksheet.cell(row=index, column=2).value = inventory_id
                except Exception as error:
                    logging.warning('  Could not create inventory for '+ item_num)
                    logging.warning(response)
                    index += 1
                    continue       
            else:
                logging.info('  ## Dry Run mode: no changes applied to Bricklink inventory ##')
        else:
            logging.info('Updating Inventory Item')

            try:
                details = getPartDetails(item_num, auth)
                if not args.dryrun:
                    worksheet.cell(row=index, column=24).value = details[item_num]['name']
                    worksheet.cell(row=index, column=7).value = details[item_num]['avg']
                else:
                    logging.info('  Latest average price is ' + str(details[item_num]['avg']))
                    if details[item_num]['avg'] > inventory_item['unit_price']:
                        logging.info('  Price has increased')
                    elif details[item_num]['avg'] < inventory_item['unit_price']:
                        logging.info('  Price has decreased')
                    else:
                        logging.info('  Price has not changed')

            except Exception as e:
                logging.warning('  Could not get pricing details for ' + str(item_num))
                logging.warning(details)
                index += 1
                continue
            # Get current online inventory quantities
            curr = get_inventory(inventory_id, auth=auth)
            logging.debug(curr)
            curr_quantity = curr['data']['quantity']

            # Update new quantity
            if curr_quantity > quantity:
                logging.info('  Reduce quantity in BL' + str(curr_quantity - quantity))
                inventory_item['quantity'] = curr_quantity - quantity
            if curr_quantity < quantity:
                logging.info('  Increase quantity in BL by ' + str(quantity - curr_quantity))
                inventory_item['quantity'] = quantity - curr_quantity
            else:
                logging.info('  No change in quantity')

            if args.dryrun:
                #logging.info(type(curr_quantity))
                #logging.info(type(quantity))
                logging.info('  Bricklink quantity: ' + str(curr_quantity))
                logging.info('  Spreadsheet quantity: ' + str(quantity))

            logging.debug(inventory_item)
            if not args.dryrun:
                response = update_inventory(inventory_id, inventory_item, auth=auth)
                logging.debug(response)
            else:
                logging.info('  ## Dry Run mode: no changes applied to Bricklink inventory ##')

        index += 1
    if not args.dryrun:
        workbook.save(filename='LegoParts.xlsx')


if __name__ == '__main__':
    main()